"""
CVRadar
Reporting Engine
src/reporting.py

v3: Clean 12-column Excel output, no internal IDs, full filenames.
Excel columns (in order):
  1  File Name         (full, no truncation)
  2  Candidate Name
  3  Primary Domain    (stream)
  4  Experience (Yrs)
  5  Education
  6  Skills Matched
  7  Skills Missing
  8  Projects
  9  Final Score
  10 Fit Category
  11 Processing Track  (Track 1 / Track 2 / Hybrid-Culled / Degraded)
  12 Filter Note       (blank unless filter triggered)
"""

from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Fit category sort order for ranking
_FIT_ORDER = {
    "Exceptional Fit": 0,
    "Strong Fit": 1,
    "Moderate Fit": 2,
    "Potential Fit": 3,
    "Low Fit": 4,
}


class ReportingEngine:

    # =====================================================
    # BUILD DATAFRAME
    # =====================================================

    @staticmethod
    def build_dataframe(candidate_profiles: list) -> pd.DataFrame:

        records = []

        for p in candidate_profiles:

            # Projects: top 3, semicolon separated
            projects_str = "; ".join(p.projects[:3]) if p.projects else ""

            # Filter note: only when filter actually triggered
            filter_note = ""
            if p.processing_notes and "FILTER:" in p.processing_notes:
                for line in p.processing_notes.splitlines():
                    if "FILTER:" in line:
                        filter_note = line.replace("FILTER:", "").strip()
                        break

            records.append({
                "File Name": p.file_name,          # full, no truncation
                "Candidate Name": p.candidate_name,
                "Primary Domain": p.primary_domain,
                "Experience (Yrs)": p.experience_years,
                "Education": p.education_summary,
                "Skills Matched": "; ".join(p.matched_requirements),
                "Skills Missing": "; ".join(p.missing_requirements),
                "Projects": projects_str,
                "Final Score": p.score_breakdown.final_score,
                "Fit Category": p.fit_category,
                "Processing Track": p.processing_track,
                "Filter Note": filter_note,
            })

        return pd.DataFrame(records)

    # =====================================================
    # RANK CANDIDATES
    # =====================================================

    @staticmethod
    def rank_candidates(df: pd.DataFrame) -> pd.DataFrame:

        if df.empty:
            return df

        df = df.sort_values(
            by="Final Score",
            ascending=False,
        ).reset_index(drop=True)

        df.insert(0, "Rank", range(1, len(df) + 1))

        return df

    # =====================================================
    # EXECUTIVE SUMMARY (for analytics cards)
    # =====================================================

    @staticmethod
    def executive_summary(df: pd.DataFrame) -> dict:

        if df.empty:
            return {}

        return {
            "total_candidates": len(df),
            "average_score": round(df["Final Score"].mean(), 1),
            "highest_score": int(df["Final Score"].max()),
            "lowest_score": int(df["Final Score"].min()),
            "shortlisted": len(df[df["Final Score"] >= 80]),
            "review": len(
                df[(df["Final Score"] >= 65) & (df["Final Score"] < 80)]
            ),
            "rejected": len(df[df["Final Score"] < 65]),
        }

    # =====================================================
    # SHORTLIST
    # =====================================================

    @staticmethod
    def shortlist(df: pd.DataFrame, threshold: int = 80) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()
        return df[df["Final Score"] >= threshold].copy()

    # =====================================================
    # EXPORT EXCEL (styled, clean, recruiter-ready)
    # =====================================================

    @staticmethod
    def export_excel(df: pd.DataFrame) -> bytes:

        output = BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:

            df.to_excel(
                writer,
                sheet_name="CVRadar_Results",
                index=False,
            )

            ws = writer.sheets["CVRadar_Results"]

            # Header styling: dark navy fill, white bold text
            header_fill = PatternFill(
                start_color="1a1a2e",
                end_color="1a1a2e",
                fill_type="solid",
            )
            header_font = Font(
                bold=True,
                color="FFFFFF",
                size=11,
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

            # Fit category colour coding
            _cat_colours = {
                "Exceptional Fit": "27AE60",  # green
                "Strong Fit": "2ECC71",        # light green
                "Moderate Fit": "F39C12",      # amber
                "Potential Fit": "E67E22",     # orange
                "Low Fit": "E74C3C",           # red
            }

            # Find column index for Fit Category
            fit_col = None
            score_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "Fit Category":
                    fit_col = idx
                if cell.value == "Final Score":
                    score_col = idx

            for row in ws.iter_rows(min_row=2):

                if fit_col:
                    cat_cell = row[fit_col - 1]
                    colour = _cat_colours.get(str(cat_cell.value), None)
                    if colour:
                        cat_cell.fill = PatternFill(
                            start_color=colour,
                            end_color=colour,
                            fill_type="solid",
                        )
                        cat_cell.font = Font(bold=True, color="FFFFFF")

                if score_col:
                    score_cell = row[score_col - 1]
                    score_cell.alignment = Alignment(horizontal="center")

            # Column widths (content-aware, capped)
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                col_name = col[0].value or ""

                for cell in col:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass

                # File Name gets extra width since we show full name
                if col_name == "File Name":
                    ws.column_dimensions[col_letter].width = min(
                        max_len + 4, 80
                    )
                elif col_name in ("Skills Matched", "Skills Missing", "Projects"):
                    ws.column_dimensions[col_letter].width = min(
                        max_len + 2, 60
                    )
                else:
                    ws.column_dimensions[col_letter].width = min(
                        max_len + 2, 40
                    )

            # Freeze top row
            ws.freeze_panes = "A2"

        output.seek(0)
        return output.getvalue()

    # =====================================================
    # MERGE BATCH REPORTS (Batch Analytics page)
    # =====================================================

    @staticmethod
    def merge_batch_reports(uploaded_files: list) -> pd.DataFrame:

        frames = []

        for f in uploaded_files:
            try:
                df = pd.read_excel(f, sheet_name="CVRadar_Results")
                frames.append(df)
            except Exception:
                try:
                    # Fallback for older Nexus reports
                    df = pd.read_excel(f, sheet_name="Candidate_Evaluation")
                    frames.append(df)
                except Exception:
                    pass

        if not frames:
            return pd.DataFrame()

        merged = pd.concat(frames, ignore_index=True)

        # Normalise column names from older exports
        rename_map = {
            "Experience": "Experience (Yrs)",
            "Skills": "Skills Matched",
            "Missing": "Skills Missing",
        }
        merged.rename(columns=rename_map, inplace=True)

        return merged
